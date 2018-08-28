import pandas as pd
from pandas_datareader import wb
import requests
import itertools
import json

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, **to_excel_kwargs):

	from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
	if 'engine' in to_excel_kwargs:
		to_excel_kwargs.pop('engine')

	writer = pd.ExcelWriter(filename, engine='openpyxl')

	try:
        # try to open an existing workbook
		writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
		if not startrow and sheet_name in writer.book.sheetnames:
			startrow = writer.book[sheet_name].max_row

        # copy existing sheets
		writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
	except FileNotFoundError:
        # file does not exist yet
		pass

	if not startrow:
		startrow = 0

	df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
	writer.save()

myKey = #input your Scopus Key
headers = {'accept':'application/json', 'x-els-apikey':myKey}
url = 'http://api.elsevier.com/content/author/author_id/'
dataFile = #input path to file you want to write data to

#get country data from pandas_datareader (World Bank data)
country_data = wb.get_countries().fillna('none')
writer = pd.ExcelWriter('country_data.xlsx')
country_data.to_excel(writer, 'Country Data')
writer.save()

#get affiliation data from Scopus, assumes you have a list of eids in a file with an 'eid' column header
df = pd.read_excel('scopus_author_data.xlsx', encoding = 'ISO-8859-1')

for row in df.itertuples():
	eid = str(row.eid)
	resp = requests.get(url + eid, headers = headers)
	auDat = resp.json()	
	columns = ['EID', 'First Name', 'Last Name', 'Initials', 'Whole Name', 'Current Affiliation', 'Affiliation History']
	if resp.status_code == 200:
		try:
			fName = auDat['author-retrieval-response'][0]['author-profile']['preferred-name']['given-name']
			lName = auDat['author-retrieval-response'][0]['author-profile']['preferred-name']['surname']
			initials = auDat['author-retrieval-response'][0]['author-profile']['preferred-name']['initials']
			indName = auDat['author-retrieval-response'][0]['author-profile']['preferred-name']['indexed-name']
			orgData = []
			affCur = auDat['author-retrieval-response'][0]['author-profile']['affiliation-current']['affiliation']['ip-doc'].get('afdispname', 'no listed current affiliation') 
			affName = auDat['author-retrieval-response'][0]['author-profile']['affiliation-history']['affiliation']
			j = 0
			while j < len(affName):
				orgAddress = affName[j]['ip-doc']['address'].setdefault('address-part', 'none')
				orgCity = affName[j]['ip-doc']['address'].setdefault('city', 'none')
				orgAbb = affName[j]['ip-doc'].setdefault('afdispname', 'none')
				orgName = affName[j]['ip-doc']['preferred-name'].setdefault('$', 'none')
				sortName = affName[j]['ip-doc'].setdefault('sort-name', 'none')
				orgState = affName[j]['ip-doc']['address'].setdefault('state', 'none')
				orgCount = affName[j]['ip-doc']['address'].setdefault('country', 'none')
				orgZip = affName[j]['ip-doc']['address'].setdefault('postal-code', 'none')
				orgId = affName[j]['ip-doc'].setdefault('@id', 'none')
				orgData.append([orgAbb, orgName, sortName, orgAddress, orgCity, orgCount, orgState, orgZip, orgId][:])
				j += 1
		except (KeyError, AttributeError, TypeError): 
			pass
	else:
		pass
	orgData = [list(map(lambda orgData: orgData + ' | ', org)) for org in orgData] #adding delimiters
	data = pd.DataFrame({'EID': [eid], 'First Name': [fName], 'Last Name':[lName], 'Initials':[initials], 'Whole Name': [indName], 'Current Affiliation':[affCur]})
	data = data.reindex(columns = columns)
	data = data.astype('object')
	data.at[0, 'Affiliation History'] = orgData
	append_df_to_excel(dataFile, data)
